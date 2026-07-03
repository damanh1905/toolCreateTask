"""Doc file Excel va sinh file Excel mau (template) cho viec import task/subtask len Jira."""
from __future__ import annotations

import datetime as _dt
import json
import os
from dataclasses import dataclass
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_NAME = "Issues"

STANDARD_COLUMNS = [
    ("LocalId", 10, "Ma tam de subtask tham chieu task cha trong cung file (VD: T1)"),
    ("IssueType", 14, "Loai issue: Task hoac Sub-task"),
    ("ProjectKey", 14, "Ma project (VD: MYTVB2C)"),
    ("ParentKey", 14, "Voi Sub-task: key Jira hoac Ma tam"),
    ("Summary", 40, "Tieu de issue (bat buoc)"),
    ("Description", 45, "Mo ta chi tiet"),
    ("Assignee", 16, "Username nguoi duoc giao"),
    ("DueDate", 14, "Han hoan thanh dinh dang YYYY-MM-DD"),
    ("TargetStart", 16, "Ngay bat dau, map vao JIRA_FIELD_TARGET_START"),
    ("TargetEnd", 16, "Ngay ket thuc, map vao JIRA_FIELD_TARGET_END"),
    ("Note", 45, "Ghi chu, map vao JIRA_FIELD_NOTE"),
    ("CustomFields", 30, "JSON cho field tuy chinh"),
]

# Cot giong form create_and_update_issue trong du an
SIMPLE_COLUMNS = [
    ("Mã", 10, "Ma tam cho task cha trong file (VD: T1). Chi can dien o dong Task"),
    ("Loại", 14, "Task hoac Sub-task"),
    ("Task cha", 18, "Voi Sub-task: nhap Ma tam hoac key Jira"),
    ("Tiêu đề", 45, "title/summary - bat buoc"),
    ("Người xử lý", 20, "assignee - username Jira, bat buoc"),
    ("Mô tả", 45, "description"),
    ("Ghi chú", 45, "note - map vao JIRA_FIELD_NOTE"),
    ("Ngày bắt đầu", 16, "start_date - map vao JIRA_FIELD_TARGET_START"),
    ("Ngày kết thúc", 16, "end_date - map vao JIRA_FIELD_TARGET_END"),
    ("Deadline", 16, "deadline/duedate"),
    ("Người tạo", 20, "Cot theo doi noi bo, tool khong gui len Jira"),
    ("Ngày tạo", 16, "Cot theo doi noi bo, tool khong gui len Jira"),
    ("Ngày update", 16, "Cot theo doi noi bo, tool khong gui len Jira"),
    ("Trạng thái", 18, "Cot theo doi noi bo, tool khong gui len Jira"),
    ("Thao tác", 16, "Cot theo doi noi bo, tool khong gui len Jira"),
]

HEADERS = [c[0] for c in STANDARD_COLUMNS]
SIMPLE_HEADERS = [c[0] for c in SIMPLE_COLUMNS]


@dataclass
class IssueRow:
    row_number: int
    local_id: str = ""
    issue_type: str = ""
    project_key: str = ""
    parent_key: str = ""
    summary: str = ""
    description: str = ""
    assignee: str = ""
    due_date: str = ""
    target_start: str = ""
    target_end: str = ""
    note: str = ""
    labels: str = ""
    custom_fields: str = ""

    @property
    def normalized_type(self) -> str:
        return self.issue_type.strip().lower().replace("-", "").replace(" ", "").replace("_", "")

    @property
    def is_subtask(self) -> bool:
        return self.normalized_type in ("subtask", "subtaskissue")

    @property
    def is_task(self) -> bool:
        return not self.is_subtask


def _is_empty_row(values: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_issues(path: str) -> list[IssueRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = [(_cell_to_str(c) or "") for c in rows[0]]
    idx = {name: header_row.index(name) for name in (HEADERS + SIMPLE_HEADERS) if name in header_row}

    def get(values, *names):
        for name in names:
            i = idx.get(name)
            if i is not None and i < len(values):
                value = _cell_to_str(values[i])
                if value:
                    return value
        return ""

    issues: list[IssueRow] = []
    for r_num, values in enumerate(rows[1:], start=2):
        values = list(values)
        if _is_empty_row(values):
            continue
        issues.append(
            IssueRow(
                row_number=r_num,
                local_id=get(values, "LocalId", "Mã"),
                issue_type=get(values, "IssueType", "Loại") or os.getenv("JIRA_DEFAULT_ISSUE_TYPE", "Task"),
                project_key=get(values, "ProjectKey") or os.getenv("JIRA_DEFAULT_PROJECT_KEY", ""),
                parent_key=get(values, "ParentKey", "Task cha"),
                summary=get(values, "Summary", "Tiêu đề"),
                description=get(values, "Description", "Mô tả"),
                assignee=get(values, "Assignee", "Người xử lý"),
                due_date=get(values, "DueDate", "Deadline"),
                target_start=get(values, "TargetStart", "Ngày bắt đầu"),
                target_end=get(values, "TargetEnd", "Ngày kết thúc"),
                note=get(values, "Note", "Ghi chú"),
                labels=get(values, "Labels"),
                custom_fields=get(values, "CustomFields"),
            )
        )
    return issues


def _resolve_labels(row: IssueRow) -> Optional[list[str]]:
    """Giong logic create_and_update_issue trong du an."""
    explicit = _split_list(row.labels)
    if explicit:
        return [label.replace(" ", "_") for label in explicit]

    if row.is_task:
        return []

    return None


def build_fields(row: IssueRow, project_key: str, parent_key: Optional[str]) -> dict:
    fields: dict[str, Any] = {
        "project": {"key": project_key},
        "summary": row.summary.strip(),
        "issuetype": {"name": _normalize_issue_type(row.issue_type)},
        "assignee": {"name": row.assignee.strip()},
    }

    if row.description:
        fields["description"] = row.description

    if parent_key:
        fields["parent"] = {"key": parent_key}

    labels = _resolve_labels(row)
    if labels is not None:
        fields["labels"] = labels

    if row.due_date:
        fields["duedate"] = row.due_date.strip()

    _add_env_custom_field(fields, "JIRA_FIELD_TARGET_START", row.target_start)
    _add_env_custom_field(fields, "JIRA_FIELD_TARGET_END", row.target_end)

    if row.custom_fields:
        try:
            extra = json.loads(row.custom_fields)
            if isinstance(extra, dict):
                fields.update(extra)
        except json.JSONDecodeError as exc:
            raise ValueError(f"CustomFields khong phai JSON hop le: {exc}") from exc

    return fields


def build_note_update(row: IssueRow) -> dict[str, Any]:
    field_key = os.getenv("JIRA_FIELD_NOTE", "").strip()
    if field_key and row.note:
        return {field_key: row.note.strip()}
    return {}


def _normalize_issue_type(value: str) -> str:
    v = value.strip()
    low = v.lower().replace("-", "").replace(" ", "")
    if low in ("subtask", "subtaskissue"):
        return os.getenv("JIRA_SUBTASK_ISSUE_TYPE", "Sub-Task").strip() or "Sub-Task"
    if low == "task":
        return "Task"
    return v


def _add_env_custom_field(fields: dict[str, Any], env_name: str, value: str) -> None:
    field_key = os.getenv(env_name, "").strip()
    if field_key and value:
        fields[field_key] = value.strip()


def _split_list(value: str) -> list[str]:
    if not value:
        return []
    return [p.strip() for p in value.split(",") if p.strip()]


def write_template(path: str, with_sample: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    note_font = Font(italic=True, color="808080", size=9)

    for col_idx, (name, width, _desc) in enumerate(SIMPLE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    if with_sample:
        samples = [
            ["T1", "Task", "", "Bổ sung suffix env vào version trên HTML5", "thanhvinh", "Mô tả task", "Ghi chú nếu có", "2026-07-01", "2026-07-10", "2026-07-15", "Trần Nguyên Phi", "2026-06-30", "", "Đang xử lý", ""],
            ["", "Sub-task", "T1", "Thiết kế UI", "thanhvinh", "Mô tả subtask", "", "2026-07-01", "2026-07-08", "2026-07-10", "Trần Nguyên Phi", "2026-06-30", "", "Đang xử lý", ""],
            ["", "Sub-task", "MYTVB2C-50784", "Gắn vào task đã tồn tại", "thanhvinh", "Mô tả subtask", "", "", "", "2026-07-15", "Trần Nguyên Phi", "2026-06-30", "", "Đang xử lý", ""],
        ]
        for r_idx, sample in enumerate(samples, start=2):
            for c_idx, val in enumerate(sample, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    ws_help = wb.create_sheet("HuongDan")
    ws_help.column_dimensions["A"].width = 20
    ws_help.column_dimensions["B"].width = 90
    ws_help.cell(row=1, column=1, value="Cot").font = Font(bold=True)
    ws_help.cell(row=1, column=2, value="Y nghia").font = Font(bold=True)
    for i, (name, _w, desc) in enumerate(SIMPLE_COLUMNS, start=2):
        ws_help.cell(row=i, column=1, value=name).font = Font(bold=True)
        c = ws_help.cell(row=i, column=2, value=desc)
        c.font = note_font
        c.alignment = Alignment(wrap_text=True, vertical="top")

    dv_type = DataValidation(type="list", formula1='"Task,Sub-task"', allow_blank=True)
    ws.add_data_validation(dv_type)
    type_col = get_column_letter(SIMPLE_HEADERS.index("Loại") + 1)
    dv_type.add(f"{type_col}2:{type_col}1000")

    wb.save(path)
